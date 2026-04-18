"""Admin contacts (authorized users) API.

Adapted from HRDDHelper/src/backend/api/v1/admin/contacts.py. Global + per-frontend
(replace|append). xlsx + csv import/export. Additive merge on import — never
destructive.
"""
import csv
import io
import logging
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from src.api.v1.admin.auth import require_admin
from src.services import contacts_store
from src.services.contacts_store import Contact, CONTACT_FIELDS, _normalise_contact

logger = logging.getLogger("admin.contacts")

router = APIRouter(prefix="/admin/api/v1/contacts", tags=["admin-contacts"])


class GlobalContactsRequest(BaseModel):
    contacts: list[Contact]


class FrontendOverrideRequest(BaseModel):
    mode: Literal["replace", "append"] = "replace"
    contacts: list[Contact] = Field(default_factory=list)


# --- Read ---

@router.get("")
async def get_contacts(_admin: dict = Depends(require_admin)):
    return contacts_store.load()


# --- Global writes ---

@router.put("/global")
async def put_global(req: GlobalContactsRequest, _admin: dict = Depends(require_admin)):
    store = contacts_store.load()
    store["global"] = [c.model_dump() for c in req.contacts]
    clean = contacts_store.save(store)
    return {"global": clean["global"]}


# --- Per-frontend overrides ---

@router.put("/frontend/{frontend_id}")
async def put_frontend(frontend_id: str, req: FrontendOverrideRequest, _admin: dict = Depends(require_admin)):
    store = contacts_store.load()
    store.setdefault("per_frontend", {})[frontend_id] = {
        "mode": req.mode,
        "contacts": [c.model_dump() for c in req.contacts],
    }
    clean = contacts_store.save(store)
    return {"frontend_id": frontend_id, "override": clean["per_frontend"].get(frontend_id)}


@router.delete("/frontend/{frontend_id}")
async def delete_frontend(frontend_id: str, _admin: dict = Depends(require_admin)):
    store = contacts_store.load()
    pf = store.get("per_frontend") or {}
    if frontend_id in pf:
        del pf[frontend_id]
        store["per_frontend"] = pf
        contacts_store.save(store)
        logger.info(f"Contacts override removed for frontend {frontend_id}")
    return {"frontend_id": frontend_id, "removed": True}


@router.post("/frontend/{frontend_id}/copy-from/{src_frontend_id}")
async def copy_from(
    frontend_id: str,
    src_frontend_id: str,
    mode: Literal["replace", "append"] = Query("replace"),
    _admin: dict = Depends(require_admin),
):
    store = contacts_store.load()
    src = (store.get("per_frontend") or {}).get(src_frontend_id)
    if not src:
        raise HTTPException(404, f"Source frontend {src_frontend_id!r} has no contacts override")
    store.setdefault("per_frontend", {})[frontend_id] = {
        "mode": mode,
        "contacts": list(src.get("contacts", [])),
    }
    clean = contacts_store.save(store)
    return {"frontend_id": frontend_id, "override": clean["per_frontend"].get(frontend_id)}


# --- Export (.xlsx) ---

@router.get("/export")
async def export_contacts(
    scope: str = Query("global", description="global | frontend:<fid> | all"),
    _admin: dict = Depends(require_admin),
):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    store = contacts_store.load()
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    def _write_sheet(sheet, contacts: list[dict[str, str]]):
        sheet.append(list(CONTACT_FIELDS))
        for c in contacts:
            sheet.append([c.get(f, "") for f in CONTACT_FIELDS])

    if scope == "all":
        ws.title = "global"
        _write_sheet(ws, store.get("global", []))
        for fid, override in (store.get("per_frontend") or {}).items():
            sheet_name = f"frontend_{fid}"[:31].replace(":", "_").replace("/", "_")
            s = wb.create_sheet(sheet_name)
            _write_sheet(s, override.get("contacts", []))
    else:
        contacts = contacts_store.contacts_for_scope(store, scope)
        ws.title = scope.replace(":", "_")[:31] or "contacts"
        _write_sheet(ws, contacts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"contacts_{scope.replace(':', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Import (.xlsx or .csv) ---

def _parse_xlsx(data: bytes) -> tuple[list[dict[str, str]], int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    if ws is None:
        return [], 0
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    out: list[dict[str, str]] = []
    ignored = 0
    for row in rows[1:]:
        raw = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        c = _normalise_contact(raw)
        if c:
            out.append(c)
        elif any(v not in (None, "") for v in row):
            ignored += 1
    return out, ignored


def _parse_csv(data: bytes) -> tuple[list[dict[str, str]], int]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    out: list[dict[str, str]] = []
    ignored = 0
    for raw in reader:
        norm_raw = {(k or "").strip().lower(): v for k, v in raw.items()}
        c = _normalise_contact(norm_raw)
        if c:
            out.append(c)
        elif any((v or "").strip() for v in raw.values()):
            ignored += 1
    return out, ignored


@router.post("/import")
async def import_contacts(
    file: UploadFile = File(...),
    scope: str = Query("global"),
    _admin: dict = Depends(require_admin),
):
    """Additive merge import from .xlsx or .csv.

    - Existing emails: fields updated where the incoming value is non-empty
    - New emails: added
    - Emails in backend but NOT in file: preserved (never deleted)
    """
    data = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        incoming, ignored = _parse_xlsx(data)
    elif name.endswith(".csv"):
        incoming, ignored = _parse_csv(data)
    else:
        raise HTTPException(400, "File must be .xlsx or .csv")

    store = contacts_store.load()
    if scope == "global":
        target = store.setdefault("global", [])
    elif scope.startswith("frontend:"):
        fid = scope.split(":", 1)[1]
        pf = store.setdefault("per_frontend", {})
        override = pf.setdefault(fid, {"mode": "replace", "contacts": []})
        target = override.setdefault("contacts", [])
    else:
        raise HTTPException(400, f"Invalid scope: {scope}")

    by_email: dict[str, dict[str, str]] = {c["email"]: c for c in target}
    added = 0
    updated = 0
    for new_c in incoming:
        existing = by_email.get(new_c["email"])
        if existing is None:
            by_email[new_c["email"]] = new_c
            added += 1
        else:
            changed = False
            for f in CONTACT_FIELDS:
                if f == "email":
                    continue
                if new_c.get(f) and new_c[f] != existing.get(f, ""):
                    existing[f] = new_c[f]
                    changed = True
            if changed:
                updated += 1

    rebuilt = list(by_email.values())
    if scope == "global":
        store["global"] = rebuilt
    else:
        fid = scope.split(":", 1)[1]
        store["per_frontend"][fid]["contacts"] = rebuilt

    contacts_store.save(store)
    logger.info(f"Contacts import ({scope}): added={added} updated={updated} ignored={ignored}")
    return {"added": added, "updated": updated, "ignored_malformed": ignored, "scope": scope}
